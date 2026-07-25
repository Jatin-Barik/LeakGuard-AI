"""Safe, local extraction helpers for statement uploads.

No uploaded document is persisted here; callers pass extracted text directly to
the transaction parser. OCR is used only for a PDF with no embedded text.
"""

from __future__ import annotations

import io

from fastapi import HTTPException


def extract_upload_text(content: bytes, filename: str) -> str:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension in {"csv", "txt", "json", "eml", "mbox"}:
        return content.decode("utf-8-sig")
    if extension in {"xlsx", "xls"}:
        return _extract_workbook(content)
    if extension == "pdf":
        return _extract_pdf(content)
    raise HTTPException(status_code=415, detail="Unsupported file type")


def _extract_workbook(content: bytes) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() if value is not None else "" for value in row]
                if any(values):
                    rows.append(",".join(values))
        return "\n".join(rows)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {error}") from error


def _extract_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader

        text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(content)).pages)
        if text.strip():
            return text
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Could not read PDF: {error}") from error

    try:
        from pdf2image import convert_from_bytes
        import pytesseract

        pages = convert_from_bytes(content, dpi=200, first_page=1, last_page=10)
        return "\n".join(pytesseract.image_to_string(page) for page in pages)
    except Exception as error:
        raise HTTPException(
            status_code=422,
            detail="This scanned PDF needs OCR support. Install Tesseract and pdf2image on the API host.",
        ) from error
